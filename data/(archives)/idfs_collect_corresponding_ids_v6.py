"""
Script to parse XLSX files in archaeometry directory and collect corresponding identifiers.

Outputs an Excel workbook with:
- Worksheet 'data'          : consolidated table built the same way but WITHOUT n_analyse (IDFS-focused)
- Worksheet 'data_ana'      : consolidated table including n_analyse

Key changes vs v3:
- All TARGET_COLUMNS are normalized and written as TEXT (strings).
- Any comma ',' found in TARGET_COLUMNS values is replaced by a dot '.'
  (to avoid comma separators / locale decimal commas).

Run as (WIndows PowerShell):
    cd data/archaeometry ; python idfs_collect_corresponding_ids_v5.py
"""

import sys
from pathlib import Path

import openpyxl
import pandas as pd

# Define the base directory
BASE_DIR = Path(__file__).parent

SUBDIRS = ['icp', 'lia', 'metallo', 'pixe', 'raman', 'sem_eds', 'tomo', 'xrf', 'xr']
TARGET_COLUMNS = ['n_iti', 'n_man', 'n_c2rmf', 'n_mrt', 'n_re', 'n_analyse']
TARGET_COLUMNS_IDFS = [c for c in TARGET_COLUMNS if c != 'n_analyse']


def find_data_files():
    """Find all data_*.xlsx files in subdirectories"""
    data_files = []
    for subdir in SUBDIRS:
        subdir_path = BASE_DIR / subdir
        if subdir_path.exists():
            for xlsx_file in subdir_path.glob('data_*.xlsx'):
                data_files.append((xlsx_file, subdir))
    return data_files


def normalize_identifier_value(val):
    """Normalize identifier values to text.

    - Always returns a string (or None).
    - Replaces any ',' by '.'
    - Strips surrounding whitespace.

    Notes:
    - If source Excel cells are numeric, Excel/OpenPyXL will already have coerced them;
      we convert to a non-scientific string where possible.
    """
    if val is None or pd.isna(val):
        return None

    if isinstance(val, bool):
        s = 'TRUE' if val else 'FALSE'
    elif isinstance(val, int):
        s = str(val)
    elif isinstance(val, float):
        if val.is_integer():
            s = str(int(val))
        else:
            s = format(val, 'f').rstrip('0').rstrip('.')
    else:
        s = str(val)

    s = s.strip().replace(',', '.')
    if s == '' or s.lower() in {'nan', '<na>'}:
        return None
    return s


def read_xlsx_data(file_path, target_columns):
    """Read data from XLSX file and extract target columns"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if 'data' not in wb.sheetnames:
            print(f"Warning: No 'data' worksheet in {file_path}", file=sys.stderr)
            return None

        ws = wb['data']

        # Get header row
        headers = [cell.value for cell in ws[1]]

        # Find indices of target columns
        col_indices = {}
        for target_col in target_columns:
            for idx, header in enumerate(headers, 1):
                if header == target_col:
                    col_indices[target_col] = idx
                    break

        if not col_indices:
            print(f"Warning: No target columns found in {file_path}", file=sys.stderr)
            return None

        # Extract data rows
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            has_data = False
            for col_name, col_idx in col_indices.items():
                cell_value = row[col_idx - 1]
                row_data[col_name] = cell_value
                if cell_value is not None and str(cell_value).strip() != '':
                    has_data = True
            if has_data:
                data_rows.append(row_data)

        return data_rows

    except Exception as e:
        print(f"Error reading {file_path}: {e}", file=sys.stderr)
        return None


def collect_all_identifiers(target_columns):
    """Collect all identifiers from all data files, tracking their sources"""
    all_rows_with_source = []  # List of (row_data, source_subdir)

    data_files = find_data_files()
    print(f"Found {len(data_files)} data files", file=sys.stderr)

    for file_path, subdir in data_files:
        print(f"Processing {file_path} (source: {subdir})", file=sys.stderr)
        rows = read_xlsx_data(file_path, target_columns=target_columns)
        if rows:
            for row in rows:
                all_rows_with_source.append((row, subdir))

    return all_rows_with_source


def create_consolidated_table(all_rows_with_source, target_columns, key_columns):
    """Create a consolidated table with mismatch detection and source tracking.

    The uniqueness key is built only from `target_columns`. If n_analyse is excluded,
    more rows will collapse into a single combination, reducing total output rows.
    """

    unique_combinations = {}  # key -> (row_data, set(source_subdirs))

    for row, source_subdir in all_rows_with_source:
        key_parts = []

        for col in key_columns:
            val = normalize_identifier_value(row.get(col))
            if val:
                key_parts.append(f"{col}:{val}")

        if not key_parts:
            continue

        key = "|".join(key_parts)

        if key not in unique_combinations:
            unique_combinations[key] = (row.copy(), set())
        else:
            existing_row, _ = unique_combinations[key]

            # Merge: fill missing values only
            for col in target_columns:
                if existing_row.get(col) is None and row.get(col) is not None:
                    existing_row[col] = row.get(col)

        unique_combinations[key][1].add(source_subdir)

    consolidated_rows = []
    all_rows_list = [row for row, _ in all_rows_with_source]

    for _, (row, source_subdirs) in unique_combinations.items():
        row_copy = row.copy()

        # Add source counts
        for subdir in SUBDIRS:
            row_copy[subdir] = 1 if subdir in source_subdirs else 0

        consolidated_rows.append(row_copy)

    return consolidated_rows


def _finalize_dataframe(consolidated_rows, target_columns):
    """Build a DataFrame with normalized text target columns and consistent ordering."""
    df = pd.DataFrame(consolidated_rows)

    # Ensure all target columns exist + normalize to TEXT
    for col in target_columns:
        if col not in df.columns:
            df[col] = None
        df[col] = (
            df[col]
            .map(normalize_identifier_value)
            .fillna('NA')
            .astype('string')
        )

    # Ensure all subdir columns exist
    for subdir in SUBDIRS:
        if subdir not in df.columns:
            df[subdir] = 0

    # Ensure description column exists
    if 'description' not in df.columns:
        df['description'] = ''

    # Reorder columns: target columns first, then subdir columns, then description
    column_order = target_columns + SUBDIRS + ['description']
    df = df[column_order]

    return df


def save_to_xlsx(df_data_ana, df_data, output_path):
    """Save both consolidated tables to a single XLSX workbook."""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_data.to_excel(writer, sheet_name='data', index=False)
        df_data_ana.to_excel(writer, sheet_name='data_ana', index=False)

    print(f"Rows in 'data': {len(df_data)}", file=sys.stderr)
    print(f"Rows in 'data_ana': {len(df_data_ana)}", file=sys.stderr)
    print(f"Consolidated tables saved to {output_path}", file=sys.stderr)


if __name__ == '__main__':
    print("Collecting identifiers from archaeometry data files...", file=sys.stderr)

    # First consolidated table (includes n_analyse)
    all_rows_with_source_full = collect_all_identifiers(target_columns=TARGET_COLUMNS)
    print(f"Total rows collected (full): {len(all_rows_with_source_full)}", file=sys.stderr)
    consolidated_rows_full = create_consolidated_table(
        all_rows_with_source_full,
        target_columns=TARGET_COLUMNS,
        key_columns=['n_iti', 'n_analyse'],
    )

    # Second consolidated table (IDFS): built the same way but WITHOUT n_analyse
    # We re-collect rows only for the columns we care about to keep logic consistent.
    all_rows_with_source_idfs = collect_all_identifiers(target_columns=TARGET_COLUMNS_IDFS)
    print(f"Total rows collected (idfs): {len(all_rows_with_source_idfs)}", file=sys.stderr)
    consolidated_rows_idfs = create_consolidated_table(
        all_rows_with_source_idfs,
        target_columns=TARGET_COLUMNS_IDFS,
        key_columns=['n_iti'],
    )

    # Build DataFrames with proper text normalization
    df_data_ana = _finalize_dataframe(consolidated_rows_full, target_columns=TARGET_COLUMNS)
    df_data = _finalize_dataframe(consolidated_rows_idfs, target_columns=TARGET_COLUMNS_IDFS)

    output_path = BASE_DIR / 'idfs_table.xlsx'
    save_to_xlsx(df_data_ana=df_data_ana, df_data=df_data, output_path=output_path)

    output_path_csv = BASE_DIR / 'idfs_table_VIEW_ONLY.csv'
    df_data.to_csv(output_path_csv, sep=',', index=False)
    
    print(f"Consolidated tables saved to {output_path_csv} (for VIEW ONLY)", file=sys.stderr)

    print("Done!", file=sys.stderr)
